# app.py
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, Response, send_from_directory
import os
import torch
from PIL import Image
from facenet_pytorch import InceptionResnetV1, MTCNN
from datetime import datetime
import openpyxl
import cv2
import numpy as np
from collections import defaultdict
import base64
from io import BytesIO


app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
UPLOAD_FOLDER = 'known_faces'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Load models
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
mtcnn = MTCNN(keep_all=False, device=device)
resnet = InceptionResnetV1(pretrained='vggface2').eval().to(device)

# Load known face embeddings
def load_known_faces():
    known_embeddings = []
    known_names = []
    for filename in os.listdir(UPLOAD_FOLDER):
        if filename.lower().endswith(('.jpg', '.png')):
            path = os.path.join(UPLOAD_FOLDER, filename)
            img = Image.open(path).convert('RGB')
            face = mtcnn(img)
            if face is not None:
                emb = resnet(face.unsqueeze(0).to(device))
                known_embeddings.append(emb.detach())
                known_names.append(os.path.splitext(filename)[0])
    return known_embeddings, known_names

known_embeddings, known_names = load_known_faces()
attendance_cache = defaultdict(bool)
current_identity = "Unknown"

# Mark attendance
def mark_attendance(name):
    filename = 'attendance.xlsx'
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['Name', 'Date', 'Time'])
        wb.save(filename)

    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    now = datetime.now()
    date, time = now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == name and row[1] == date:
            return
    sheet.append([name, date, time])
    wb.save(filename)

# Live face recognition
def gen_frames():
    global current_identity
    cap = cv2.VideoCapture(0)
    while True:
        success, frame = cap.read()
        if not success:
            break

        img = Image.fromarray(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
        face = mtcnn(img)

        identity = "Unknown"
        if face is not None:
            emb = resnet(face.unsqueeze(0).to(device))
            max_sim = 0
            for known_emb, name in zip(known_embeddings, known_names):
                sim = torch.nn.functional.cosine_similarity(emb, known_emb).item()
                if sim > max_sim and sim > 0.6:
                    max_sim = sim
                    identity = name

            if identity != "Unknown" and not attendance_cache[identity]:
                mark_attendance(identity)
                attendance_cache[identity] = True

        current_identity = identity

        ret, buffer = cv2.imencode('.jpg', frame)
        frame_bytes = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')

# Auth Middleware
@app.before_request
def require_login():
    allowed_routes = {'login', 'static', 'video_feed', 'current_name', 'register_face'}
    if request.endpoint not in allowed_routes and not session.get('admin'):
        return redirect(url_for('login'))

# Routes

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form['username'] == 'admin' and request.form['password'] == 'admin123':
            session['admin'] = True
            return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('login'))

@app.route('/view_attendance')
def view_attendance():
    filename = 'attendance.xlsx'
    if not os.path.exists(filename):
        return render_template('attendance.html', data=[])
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    data = [row for row in sheet.iter_rows(values_only=True)]
    return render_template('attendance.html', data=data)

known_embeddings, known_names = load_known_faces()

@app.route('/register', methods=['GET', 'POST'])
def register():
    global known_embeddings, known_names
    if request.method == 'POST':
        name = request.form['name'].strip()
        img_data = request.form['imageData'].split(",")[1]
        img_bytes = base64.b64decode(img_data)
        img = Image.open(BytesIO(img_bytes))
        img.save(os.path.join(UPLOAD_FOLDER, f"{name}.jpg"))

        # ⬅ Reload embeddings after registration
        known_embeddings, known_names = load_known_faces()

        return redirect(url_for('register'))

    students = []
    for f in os.listdir(UPLOAD_FOLDER):
        if f.endswith(".jpg"):
            students.append({
                'name': f.rsplit('.', 1)[0],
                'img_path': f"/known_faces/{f}"
            })
    return render_template('register.html', students=students)


@app.route('/known_faces/<filename>')
def known_faces(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)


@app.route('/delete/<name>', methods=['POST'])
def delete_student(name):
    file_path = os.path.join(UPLOAD_FOLDER, f"{name}.jpg")
    if os.path.exists(file_path):
        os.remove(file_path)
    return redirect(url_for('register'))


@app.route('/register_face', methods=['POST'])
def register_face():
    file = request.files['image']
    name = file.filename.rsplit('.', 1)[0]
    save_path = os.path.join(UPLOAD_FOLDER, f"{name}.jpg")
    file.save(save_path)
    global known_embeddings, known_names
    known_embeddings, known_names = load_known_faces()
    return "Registered Successfully!"

@app.route('/video_feed')
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/current_name')
def current_name():
    return jsonify({'name': current_identity})

@app.route('/static/known_faces/<filename>')
def static_known_faces(filename):
    return send_from_directory(UPLOAD_FOLDER, filename)

@app.route('/')
def dashboard():
    total_students = len(known_names)

    filename = 'attendance.xlsx'
    today = datetime.now().strftime("%Y-%m-%d")
    current_weekday = datetime.now().weekday()  # Monday=0 ... Sunday=6
    daily_counts = {i: 0 for i in range(6)}  # Mon-Sat only

    todays_attendance = set()
    recent_verifications = []
    new_registrations = 0
    all_dates = set()

    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active

        for row in reversed(list(sheet.iter_rows(min_row=2, values_only=True))):
            name, date, time = row
            all_dates.add(date)

            weekday = datetime.strptime(date, "%Y-%m-%d").weekday()
            if weekday < 6:
                daily_counts[weekday] += 1

            if date == today:
                todays_attendance.add(name)
                if len(recent_verifications) < 5:
                    recent_verifications.append(f"{time} - {name} marked present ✅")

    # Get new registrations today (based on file created time)
    for f in os.listdir(UPLOAD_FOLDER):
        if f.endswith(".jpg"):
            creation_time = datetime.fromtimestamp(os.path.getctime(os.path.join(UPLOAD_FOLDER, f)))
            if creation_time.strftime("%Y-%m-%d") == today:
                new_registrations += 1

    attendance_percent = round((len(todays_attendance) / total_students) * 100, 1) if total_students else 0

    late_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, date, time = row
        if date == today:
            attendance_time = datetime.strptime(time, "%H:%M:%S").time()
            if attendance_time > datetime.strptime("09:30:00", "%H:%M:%S").time():
                late_count += 1

    chart_data = {
        "labels": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"],
        "values": [daily_counts[i] for i in range(6)],
        "breakdown": [
            len(todays_attendance),                     # Present
            total_students - len(todays_attendance),    # Absent
            late_count                                  # Late
        ]
    }

    return render_template('dashboard.html',
        total_students=total_students,
        attendance_percent=attendance_percent,
        new_registrations=new_registrations,
        unrecognized_faces=0,  # Placeholder, can update later if needed
        chart_data=chart_data,
        recent_verifications=recent_verifications
    )


@app.route('/live')
def live_attendance():
    return render_template('live.html')

@app.route('/download_attendance')
def download_attendance():
    filepath = 'attendance.xlsx'
    if os.path.exists(filepath):
        return send_from_directory(directory='.', path=filepath, as_attachment=True)
    return "No attendance file found", 404

@app.route('/delete_record', methods=['POST'])
def delete_record():
    name = request.form['name']
    date = request.form['date']
    time = request.form['time']
    filename = 'attendance.xlsx'

    if not os.path.exists(filename):
        return redirect(url_for('view_attendance'))

    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    for row in list(sheet.iter_rows(min_row=2)):
        if (row[0].value == name and row[1].value == date and row[2].value == time):
            sheet.delete_rows(row[0].row, 1)
            break

    wb.save(filename)
    return redirect(url_for('view_attendance'))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))  # Default fallback
    app.run(host="0.0.0.0", port=port)

